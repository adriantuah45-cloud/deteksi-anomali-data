"""
excel_exporter.py
Menyusun DataFrame hasil pencocokan/anomali menjadi file Excel (.xlsx)
dengan 3 sheet: "Semua Data", "Anomali", dan "Ringkasan", lengkap
dengan warna status, freeze panes, dan auto filter.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="0E1B2B", end_color="0E1B2B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

STATUS_FILLS = {
    "Cocok": PatternFill(start_color="E6F7EE", end_color="E6F7EE", fill_type="solid"),
    "Anomali Ringan": PatternFill(start_color="FDF3DA", end_color="FDF3DA", fill_type="solid"),
    "Anomali Berat": PatternFill(start_color="FBE6E7", end_color="FBE6E7", fill_type="solid"),
}

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_sheet(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = BORDER

    status_col_idx = list(df.columns).index("Status") + 1 if "Status" in df.columns else None

    for row in df.itertuples(index=False):
        ws.append(list(row))
        r = ws.max_row
        status_val = row[status_col_idx - 1] if status_col_idx else None
        fill = STATUS_FILLS.get(status_val)
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if fill:
                cell.fill = fill

    for i, col in enumerate(df.columns, 1):
        col_len = df[col].astype(str).map(len).max() if len(df) else 0
        ws.column_dimensions[get_column_letter(i)].width = min(max(col_len, len(str(col))) + 4, 40)

    ws.freeze_panes = "A2"
    if len(df):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"


def build_summary_df(result_df: pd.DataFrame) -> pd.DataFrame:
    total = len(result_df)
    cocok = int((result_df["Status"] == "Cocok").sum())
    ringan = int((result_df["Status"] == "Anomali Ringan").sum())
    berat = int((result_df["Status"] == "Anomali Berat").sum())
    persen = f"{(cocok / total * 100 if total else 0):.1f}%"

    return pd.DataFrame({
        "Metrik": ["Total Data", "Data Cocok", "Anomali Ringan", "Anomali Berat", "Tingkat Kecocokan"],
        "Nilai": [total, cocok, ringan, berat, persen],
    })


def export_to_excel(result_df: pd.DataFrame) -> bytes:
    """Susun hasil ke file Excel 3 sheet dan kembalikan sebagai bytes (siap diunduh)."""
    wb = Workbook()

    ws_all = wb.active
    ws_all.title = "Semua Data"
    _write_sheet(ws_all, result_df)

    anomali_df = result_df[result_df["Status"] != "Cocok"].reset_index(drop=True)
    ws_anomali = wb.create_sheet("Anomali")
    _write_sheet(ws_anomali, anomali_df if len(anomali_df) else result_df.iloc[0:0])

    ws_summary = wb.create_sheet("Ringkasan")
    _write_sheet(ws_summary, build_summary_df(result_df))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
